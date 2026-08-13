"""
API de grupos/áreas, atribuição de trilhas, carga em lote de usuários e
gamificação.

Trainee (get_current_user):
  GET  /api/me/tracks           → trilhas + simulados visíveis + gamificação
  GET  /api/me/gamification     → pontos/nível/medalhas do usuário
  GET  /api/me/certifications   → simulados visíveis para o usuário

Admin (require_admin):
  GET/POST/PATCH/DELETE /api/admin/groups[...]   → CRUD de grupos
  PATCH  /api/admin/users/{email}/group          → vincular usuário a grupo/trilhas
  POST   /api/admin/users/bulk                    → carga em lote (CSV/XLSX)
  GET    /api/admin/users/template.(csv|xlsx)     → modelo de planilha
  GET    /api/admin/tracks/overview               → visão por trilha (matrículas/progresso)
  GET    /api/gamification/leaderboard            → ranking (por grupo, opcional)
"""
from __future__ import annotations

import io
import logging
import re
import secrets
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response, UploadFile, File, Form
from pydantic import BaseModel

from app.auth import security
from app.config import get_settings
from app.models.schemas import UserPublic
from app.services import groups as groups_svc
from app.services import gamification as gami
from app.services import users as users_svc
from app.services import repo, activity

logger = logging.getLogger(__name__)
router = APIRouter()

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
DEFAULT_BULK_PASSWORD = "Databricks#ATT2026"


# ── Modelos ───────────────────────────────────────────────────────────────────
class GroupBody(BaseModel):
    key: str = ""
    name: str = ""
    description: str = ""
    color: str = "#00A8E0"
    icon: str = ""
    track_keys: List[str] = []
    certification_ids: List[str] = []
    sort_order: int = 0


class GroupPatch(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    color: Optional[str] = None
    icon: Optional[str] = None
    track_keys: Optional[List[str]] = None
    certification_ids: Optional[List[str]] = None
    sort_order: Optional[int] = None


class UserGroupBody(BaseModel):
    group_key: Optional[str] = None
    extra_track_keys: Optional[List[str]] = None


# ── Trainee ────────────────────────────────────────────────────────────────────
@router.get("/me/tracks")
async def my_tracks(user: UserPublic = Depends(security.get_current_user)):
    vis = groups_svc.visible_tracks_for_user(user.tenant_id, user.email, is_admin=user.is_admin)
    stats = gami.user_stats(user.tenant_id, user.email)
    return {**vis, "gamification": stats}


@router.get("/me/gamification")
async def my_gamification(user: UserPublic = Depends(security.get_current_user)):
    return gami.user_stats(user.tenant_id, user.email)


@router.get("/me/certifications")
async def my_certifications(user: UserPublic = Depends(security.get_current_user)):
    """Simulados visíveis para o usuário (filtrados por grupo/trilhas)."""
    vis = groups_svc.visible_tracks_for_user(user.tenant_id, user.email, is_admin=user.is_admin)
    certs = repo.list_certifications()
    if vis.get("all_sims"):
        return certs
    allowed = set(vis.get("sim_cert_ids", []))
    return [c for c in certs if (c.get("id") if isinstance(c, dict) else getattr(c, "id", None)) in allowed]


@router.get("/groups")
async def list_areas(user: UserPublic = Depends(security.get_current_user)):
    """Áreas (grupos) com suas trilhas resolvidas — navegação por área (estilo Genie).

    Aberto a qualquer usuário logado: é um catálogo de aprendizagem, não dado sensível."""
    from app.services import tenants as tenants_svc
    tt = tenants_svc.get_tenant_by_id(user.tenant_id)
    all_tracks = (tenants_svc.get_routes(tt["slug"]) if tt else {}).get("routes", [])
    by_key = {t.get("key"): t for t in all_tracks if t.get("key")}
    areas = []
    for g in groups_svc.list_groups(user.tenant_id):
        tks = g.get("track_keys") or []
        tracks = [by_key[k] for k in tks if k in by_key] if tks else all_tracks
        areas.append({
            **g, "tracks": tracks,
            "n_tracks": len(tracks),
            "n_classes": sum(len(t.get("classes", []) or []) for t in tracks),
        })
    return {"areas": areas}


# ── Ranking (gamificação) ───────────────────────────────────────────────────────
@router.get("/gamification/leaderboard")
async def gami_leaderboard(user: UserPublic = Depends(security.get_current_user),
                           group: Optional[str] = Query(None),
                           scope: str = Query("all")):
    """Ranking por pontos. scope='group' limita ao grupo do próprio usuário."""
    gk = group
    if scope == "group" and not gk:
        gk = groups_svc.get_user_membership(user.tenant_id, user.email).get("group_key")
    rows = gami.leaderboard(user.tenant_id, group_key=gk)
    groups = groups_svc.list_groups(user.tenant_id)
    return {"rows": rows, "groups": groups, "group_key": gk}


# ── Admin: CRUD de grupos ────────────────────────────────────────────────────────
@router.get("/admin/groups")
async def admin_list_groups(admin: UserPublic = Depends(security.require_admin)):
    return groups_svc.list_groups(admin.tenant_id)


@router.post("/admin/groups")
async def admin_create_group(body: GroupBody, admin: UserPublic = Depends(security.require_admin)):
    key = re.sub(r"[^a-z0-9_]+", "_", (body.key or body.name).strip().lower()).strip("_")
    if not key:
        raise HTTPException(422, "key/nome inválido")
    if not body.name.strip():
        raise HTTPException(422, "Nome obrigatório")
    try:
        return groups_svc.create_group(
            admin.tenant_id, key, body.name.strip(), body.description, body.color,
            body.icon, body.track_keys, body.certification_ids, body.sort_order)
    except ValueError as e:
        raise HTTPException(409, str(e))


@router.patch("/admin/groups/{key}")
async def admin_update_group(key: str, body: GroupPatch,
                             admin: UserPublic = Depends(security.require_admin)):
    g = groups_svc.update_group(admin.tenant_id, key, **body.model_dump(exclude_none=True))
    if not g:
        raise HTTPException(404, "Grupo não encontrado")
    return g


@router.delete("/admin/groups/{key}")
async def admin_delete_group(key: str, admin: UserPublic = Depends(security.require_admin)):
    groups_svc.delete_group(admin.tenant_id, key)
    return {"ok": True}


# ── Admin: vínculo usuário ↔ grupo ────────────────────────────────────────────────
@router.patch("/admin/users/{email}/group")
async def admin_set_user_group(email: str, body: UserGroupBody,
                               admin: UserPublic = Depends(security.require_admin)):
    email = email.lower()
    if not users_svc.get_user(admin.tenant_id, email):
        raise HTTPException(404, "Usuário não encontrado")
    if body.group_key is not None:
        gk = body.group_key or None
        if gk and not groups_svc.get_group(admin.tenant_id, gk):
            raise HTTPException(422, f"Grupo '{gk}' inexistente")
        groups_svc.set_user_group(admin.tenant_id, email, gk)
    if body.extra_track_keys is not None:
        groups_svc.set_user_extra_tracks(admin.tenant_id, email, body.extra_track_keys)
    return {"ok": True}


# ── Admin: carga de usuários em lote ──────────────────────────────────────────────
def _parse_rows(filename: str, raw: bytes) -> list[dict]:
    """Extrai linhas (dicts com cabeçalhos normalizados) de CSV ou XLSX."""
    name = (filename or "").lower()
    headers: list[str] = []
    records: list[list] = []
    if name.endswith(".xlsx") or raw[:2] == b"PK":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = ["" if c is None else str(c).strip() for c in row]
            if i == 0:
                headers = [h.strip().lower() for h in vals]
            elif any(vals):
                records.append(vals)
    else:
        import csv
        text = raw.decode("utf-8-sig", errors="replace")
        # detecta separador (vírgula ou ponto-e-vírgula)
        sample = text.splitlines()[0] if text.splitlines() else ""
        delim = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.reader(io.StringIO(text), delimiter=delim)
        for i, row in enumerate(reader):
            vals = [c.strip() for c in row]
            if i == 0:
                headers = [h.strip().lower() for h in vals]
            elif any(vals):
                records.append(vals)

    # mapeia sinônimos de colunas → campos canônicos
    alias = {
        "name": ["nome", "name", "nombre", "nome completo"],
        "email": ["email", "e-mail", "correo", "mail"],
        "area": ["area", "área", "departamento", "sector"],
        "group": ["grupo", "group", "rota", "ruta", "rota sugerida", "ruta sugerida", "trilha", "área/grupo"],
    }
    idx = {}
    for canon, names in alias.items():
        for j, h in enumerate(headers):
            if h in names:
                idx[canon] = j
                break
    out = []
    for vals in records:
        def g(k):
            j = idx.get(k)
            return vals[j].strip() if j is not None and j < len(vals) else ""
        out.append({"name": g("name"), "email": g("email").lower(),
                    "area": g("area"), "group": g("group").lower()})
    return out


@router.post("/admin/users/bulk")
async def admin_bulk_users(request: Request,
                           file: UploadFile = File(...),
                           default_password: str = Form(DEFAULT_BULK_PASSWORD),
                           admin: UserPublic = Depends(security.require_admin)):
    """Cria usuários em lote a partir de uma planilha (CSV/XLSX).

    Colunas aceitas (cabeçalho, PT/ES/EN): nome, email, area, grupo (rota sugerida).
    O 'grupo' pode ser a KEY ou o NOME de um grupo existente. Usuários novos recebem
    a senha inicial informada (troca obrigatória no 1º acesso). E-mails já existentes
    apenas têm grupo/área atualizados."""
    raw = await file.read()
    if not raw:
        raise HTTPException(422, "Arquivo vazio")
    if len(default_password) < 6:
        raise HTTPException(422, "A senha inicial deve ter ao menos 6 caracteres")
    try:
        rows = _parse_rows(file.filename or "", raw)
    except Exception as e:
        raise HTTPException(422, f"Falha ao ler a planilha: {e}")

    # resolve grupos por key OU nome
    groups = groups_svc.list_groups(admin.tenant_id)
    by_key = {g["key"]: g["key"] for g in groups}
    by_name = {g["name"].strip().lower(): g["key"] for g in groups}

    created, updated, errors = 0, 0, []
    pw_hash = security.hash_password(default_password)
    for i, r in enumerate(rows, start=2):  # linha 1 = cabeçalho
        email, name = r["email"], r["name"]
        if not email or not _EMAIL_RE.match(email):
            errors.append({"row": i, "email": email, "error": "e-mail inválido"})
            continue
        if not name:
            name = email.split("@")[0]
        gk = None
        if r["group"]:
            gk = by_key.get(r["group"]) or by_name.get(r["group"])
            if not gk:
                errors.append({"row": i, "email": email,
                               "error": f"grupo '{r['group']}' inexistente"})
                # segue criando o usuário, apenas sem grupo
        exists = users_svc.get_user(admin.tenant_id, email)
        if exists:
            users_svc.update_profile(admin.tenant_id, email, name or None, r["area"] or None)
            if gk:
                groups_svc.set_user_group(admin.tenant_id, email, gk)
            updated += 1
        else:
            users_svc.create_user(admin.tenant_id, email, name, pw_hash,
                                  is_admin=False, must_change_password=True, area=r["area"] or None)
            if gk:
                groups_svc.set_user_group(admin.tenant_id, email, gk)
            created += 1

    activity.log_event(admin.tenant_id, admin.email, "bulk_import", request=request,
                       detail={"created": created, "updated": updated, "errors": len(errors)})
    return {"created": created, "updated": updated, "total": len(rows),
            "errors": errors, "default_password": default_password}


_TEMPLATE_ROWS = [
    ["nome", "email", "area", "grupo"],
    ["Ana López", "ana.lopez@att.com", "Oficina del CDO", "cdo"],
    ["Bruno Díaz", "bruno.diaz@att.com", "Ciberseguridad", "cso"],
    ["Carla Ruiz", "carla.ruiz@att.com", "Finanzas", "finanzas"],
]


@router.get("/admin/users/template.csv")
async def template_csv(_: UserPublic = Depends(security.require_admin)):
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)
    for row in _TEMPLATE_ROWS:
        w.writerow(row)
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": 'attachment; filename="modelo_usuarios_att.csv"'})


@router.get("/admin/users/template.xlsx")
async def template_xlsx(_: UserPublic = Depends(security.require_admin)):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuários"
    for row in _TEMPLATE_ROWS:
        ws.append(row)
    for c in ws[1]:
        from openpyxl.styles import Font
        c.font = Font(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="modelo_usuarios_att.xlsx"'})


# ── Admin: visão por trilha ──────────────────────────────────────────────────────
@router.get("/admin/tracks/overview")
async def admin_tracks_overview(admin: UserPublic = Depends(security.require_admin)):
    """Para cada trilha: usuários matriculados (por grupo) e o progresso deles.

    'Matriculado' = usuário cujo grupo inclui a trilha (ou trilha extra do usuário)."""
    tenant_id = admin.tenant_id
    from app.services import tenants as tenants_svc
    tt = tenants_svc.get_tenant_by_id(tenant_id)
    tracks = (tenants_svc.get_routes(tt["slug"]) if tt else {}).get("routes", [])
    groups = groups_svc.list_groups(tenant_id)

    # mapa grupo→trilhas e usuários→grupo
    all_users = users_svc.list_users(tenant_id)
    memberships = {u["email"]: groups_svc.get_user_membership(tenant_id, u["email"]) for u in all_users}
    grp_by_key = {g["key"]: g for g in groups}

    # progresso agregado
    from app.db import get_conn
    prog: dict[str, set] = {}          # email -> set(class_id)
    attempts: dict[tuple, dict] = {}   # (email,cert) -> {n, passed, best}
    if not get_settings().MOCK_MODE:
        with get_conn() as conn:
            for e, cid in conn.execute(
                    "SELECT user_email, class_id FROM class_progress WHERE tenant_id=%s",
                    (tenant_id,)).fetchall():
                prog.setdefault(e, set()).add(cid)
            for e, cert, passed, sc in conn.execute(
                    "SELECT user_email, certification_id, passed, score_pct FROM test_sessions "
                    "WHERE tenant_id=%s", (tenant_id,)).fetchall():
                k = (e, cert)
                a = attempts.setdefault(k, {"n": 0, "passed": 0, "best": 0.0})
                a["n"] += 1
                a["passed"] += 1 if passed else 0
                a["best"] = max(a["best"], sc or 0.0)

    out = []
    for tr in tracks:
        tkey = tr.get("key")
        class_ids = [c.get("id") for c in tr.get("classes", []) if c.get("id")]
        n_classes = len(class_ids)
        cert_ids = ([tr["certification_id"]] if tr.get("certification_id") else []) + \
                   (tr.get("sim_cert_ids") or [])
        cert_ids = list(dict.fromkeys(cert_ids))
        enrolled = []
        for u in all_users:
            m = memberships[u["email"]]
            g = grp_by_key.get(m["group_key"]) if m["group_key"] else None
            in_track = bool(g and tkey in (g.get("track_keys") or [])) or \
                       (tkey in (m.get("extra_track_keys") or []))
            if not in_track:
                continue
            done = len(prog.get(u["email"], set()) & set(class_ids))
            att = sum(attempts.get((u["email"], c), {}).get("n", 0) for c in cert_ids)
            passed = sum(attempts.get((u["email"], c), {}).get("passed", 0) for c in cert_ids)
            enrolled.append({
                "email": u["email"], "name": u["name"], "area": u.get("area"),
                "group_key": m["group_key"],
                "classes_done": done, "classes_total": n_classes,
                "pct": round(100 * done / n_classes) if n_classes else 0,
                "attempts": att, "passed": passed,
            })
        enrolled.sort(key=lambda x: (-x["pct"], -x["attempts"], x["name"] or ""))
        avg_pct = round(sum(e["pct"] for e in enrolled) / len(enrolled)) if enrolled else 0
        out.append({
            "key": tkey, "name": tr.get("name"), "icon": tr.get("icon"),
            "color": tr.get("color"), "description": tr.get("description"),
            "classes_total": n_classes, "cert_ids": cert_ids,
            "enrolled_count": len(enrolled), "avg_pct": avg_pct,
            "completed_count": sum(1 for e in enrolled if e["pct"] >= 100),
            "enrolled": enrolled,
        })
    return {"tracks": out, "groups": groups}
